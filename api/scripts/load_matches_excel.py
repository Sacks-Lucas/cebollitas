"""load_matches_excel.py — One-time loader: "Partiditos de la gente" Excel → JSON.

Reads the source Excel and rewrites api/data/matches.json, which the persistence
layer seeds into the `matches` collection (Mongo or local JSON) on boot.

Run from the api/ directory:
    venv/Scripts/python.exe scripts/load_matches_excel.py

Safe to re-run: ids are deterministic (uuid5 per row), so the file is rebuilt
identically. The Excel sheet has columns: Nombre | Fecha del partido | Resultado
| Goles | Estadio.
"""

import json
import uuid
from datetime import datetime
from pathlib import Path

API_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = API_DIR / "data"
EXCEL_PATH = Path(r"D:\Descargas\Partiditos de la gente.xlsx")
SHEET = "2026"

# Excel result (Spanish) → canonical English code stored in the collection. The
# UI translates these back to Spanish via i18n.
RESULT_BY_LABEL = {
    "victoria": "win",
    "derrota": "loss",
    "empate": "draw",
}

# Excel nickname → canonical full name in allowed_users. Each match is linked to
# the user's id (FK), resolved from allowed_users.json at load time.
NAME_BY_NICK = {
    "nacho": "Ignacio Calvo",
    "petro": "Martin Petrone",
    "char": "Carlos Martinez Devesa",
    "axel": "Axel Reynoso",
    "cala": "Santiago Cala",
    "renzo": "Renzo Tamburella",
    "luqui": "Lucas Mateo Sacks",
    "mateo": "Mateo Lozano",
}


def _user_id_by_nick() -> dict[str, str]:
    """Build nickname → userId, failing fast if a mapped name is not an allowed user."""
    users = json.loads((DATA_DIR / "allowed_users.json").read_text(encoding="utf-8"))
    id_by_name = {user["name"]: user["id"] for user in users}
    missing = sorted({name for name in NAME_BY_NICK.values() if name not in id_by_name})
    if missing:
        raise SystemExit(f"ERROR: these mapped names are not in allowed_users.json: {missing}")
    return {nick: id_by_name[name] for nick, name in NAME_BY_NICK.items()}


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _match_id(row_num: int, name: str, date_str: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"cebollitas:match:{row_num}:{name}:{date_str}"))


def main() -> None:
    if not EXCEL_PATH.exists():
        raise SystemExit(f"ERROR: Excel not found at {EXCEL_PATH}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    user_id_by_nick = _user_id_by_nick()

    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[SHEET]

    matches: list[dict] = []
    skipped = 0
    for row_num in range(2, ws.max_row + 1):
        name = _clean(ws.cell(row_num, 1).value)
        fecha = ws.cell(row_num, 2).value
        result_label = _clean(ws.cell(row_num, 3).value)
        goals_raw = ws.cell(row_num, 4).value
        stadium = _clean(ws.cell(row_num, 5).value)

        # Skip fully empty rows.
        if not any([name, fecha, result_label, goals_raw, stadium]):
            continue

        if not name or not isinstance(fecha, datetime) or not result_label:
            print(f"  SKIP row {row_num}: missing name/date/result ({name!r}, {fecha!r}, {result_label!r})")
            skipped += 1
            continue

        result = RESULT_BY_LABEL.get(result_label.lower())
        if result is None:
            print(f"  SKIP row {row_num}: unknown result {result_label!r}")
            skipped += 1
            continue

        date_str = fecha.date().isoformat()
        # Missing goals count as 0 (so stats aggregate cleanly).
        goals = int(goals_raw) if isinstance(goals_raw, (int, float)) else 0

        user_id = user_id_by_nick.get(name.lower())
        if user_id is None:
            print(f"  SKIP row {row_num}: unmapped nickname {name!r} (no matching user)")
            skipped += 1
            continue

        matches.append(
            {
                "id": _match_id(row_num, name, date_str),
                "userId": user_id,
                "date": date_str,
                "result": result,
                "goals": goals,
                "stadium": stadium,
            }
        )

    out_path = DATA_DIR / "matches.json"
    out_path.write_text(json.dumps(matches, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Done. {len(matches)} matches written to {out_path} ({skipped} skipped).")


if __name__ == "__main__":
    main()
