"""
load_excel.py — One-time historical data loader: Excel → JSON persistence files.

Run with:
    api/.venv/Scripts/python.exe api/scripts/load_excel.py

Rewrites:
    api/data/allowed_users.json
    api/data/events.json
    api/data/votes.json
    api/data/monthly_assignments.json  (fixes Mateo name)

Also ensures api/.env has VOTE_ENCRYPTION_KEY and JWT_SECRET set.
Safe to re-run: deterministic UUIDs for users; events/votes are fully replaced.

Excluded events (per user decision):
  - Juntada Express: Tipas CJ (row 4)  — to be loaded manually later
  - Vacaciones CRUCERO (row 9)          — trips handled separately
  - Viaje express ROSARIO (row 13)      — trips handled separately
  - Evento Nacho: ????? (row 25)        — future placeholder, no data
"""

import hashlib
import json
import math
import os
import secrets
import uuid
from datetime import datetime, timezone
from pathlib import Path

from cryptography.fernet import Fernet
from dotenv import load_dotenv, set_key

# ── Paths ──────────────────────────────────────────────────────────────────────
API_DIR   = Path(__file__).resolve().parent.parent
DATA_DIR  = API_DIR / "data"
ENV_FILE  = API_DIR / ".env"
EXCEL_PATH = Path(r"D:\Descargas\Copia de excel contrato falopa.xlsx")

# ── Env setup ──────────────────────────────────────────────────────────────────

def ensure_env_keys() -> str:
    """Return the VOTE_ENCRYPTION_KEY, generating it (and JWT_SECRET) if absent."""
    load_dotenv(ENV_FILE)

    if not os.getenv("VOTE_ENCRYPTION_KEY"):
        key = Fernet.generate_key().decode()
        set_key(str(ENV_FILE), "VOTE_ENCRYPTION_KEY", key)
        os.environ["VOTE_ENCRYPTION_KEY"] = key
        print(f"  Generated VOTE_ENCRYPTION_KEY -> {ENV_FILE}")
    else:
        print(f"  VOTE_ENCRYPTION_KEY already set.")

    if not os.getenv("JWT_SECRET"):
        secret = secrets.token_hex(32)
        set_key(str(ENV_FILE), "JWT_SECRET", secret)
        os.environ["JWT_SECRET"] = secret
        print(f"  Generated JWT_SECRET -> {ENV_FILE}")
    else:
        print(f"  JWT_SECRET already set.")

    return os.environ["VOTE_ENCRYPTION_KEY"]


# ── User definitions ───────────────────────────────────────────────────────────

LUCAS_EXISTING_ID = "86d65b13-6d9c-438f-a0d9-1a61c6d30e3b"

def _stable_id(name: str) -> str:
    """Deterministic UUID5 so re-runs produce the same IDs."""
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"cebollitas:{name.lower()}"))

# Nickname (Excel header) → user record
USERS_BY_NICK: dict[str, dict] = {
    "Axel":    {"name": "Axel Reynoso",          "email": "axel.reynoso@cebollitas.local"},
    "Cala":    {"name": "Santiago Cala",          "email": "santiago.cala@cebollitas.local"},
    "Nacho":   {"name": "Ignacio Calvo",          "email": "ignacio.calvo@cebollitas.local"},
    "Carlos":  {"name": "Carlos Martinez Devesa", "email": "carlos.martinez@cebollitas.local"},
    "Fer":     {"name": "Fernando Suarez",        "email": "fernando.suarez@cebollitas.local"},
    "Lucky":   {"name": "Lucas Mateo Sacks",      "email": "sackslm0@gmail.com",
                "id": LUCAS_EXISTING_ID},          # keep real UUID + real email
    "Winnie":  {"name": "Mateo Lozano",           "email": "mateo.lozano@cebollitas.local"},
    "Pela":    {"name": "Nicolas Escobar",        "email": "nicolas.escobar@cebollitas.local"},
    "Petro":   {"name": "Martin Petrone",         "email": "martin.petrone@cebollitas.local"},
    "Renzo":   {"name": "Renzo Tamburella",       "email": "renzo.tamburella@cebollitas.local"},
    "Seba":    {"name": "Sebastian Arias",        "email": "sebastian.arias@cebollitas.local"},
    "Risitos": {"name": "Ignacio Taricco",        "email": "ignacio.taricco@cebollitas.local"},
}

for _nick, _info in USERS_BY_NICK.items():
    if "id" not in _info:
        _info["id"] = _stable_id(_info["name"])

# Column order in Excel sheet (E..P)
NICK_ORDER = ["Axel", "Cala", "Nacho", "Carlos", "Fer",
              "Lucky", "Winnie", "Pela", "Petro", "Renzo", "Seba", "Risitos"]


# ── Per-row classification ─────────────────────────────────────────────────────
# Keys: skip, type, org (nick or None), vote_avg (monthly_event only)

EVENT_META: dict[int, dict] = {
    4:  {"skip": True},                              # Juntada Express: Tipas CJ (deferred)
    5:  {"type": "regular",       "org": "Petro"},
    6:  {"type": "regular",       "org": "Petro"},
    7:  {"type": "regular",       "org": "Cala"},
    8:  {"type": "monthly_event", "org": "Petro",  "vote_avg": 7.0},
    9:  {"skip": True},                              # Crucero (trip, deferred)
    10: {"type": "monthly_event", "org": "Cala",   "vote_avg": 7.0},
    11: {"type": "extended",      "org": None},      # Continuación Canton
    12: {"type": "regular",       "org": "Carlos"},
    13: {"skip": True},                              # Rosario (trip, sin asistentes)
    14: {"type": "regular",       "org": "Petro"},
    15: {"type": "monthly_event", "org": "Renzo",  "vote_avg": 8.0},
    16: {"type": "extended",      "org": None},      # Express: Pompeya
    17: {"type": "regular",       "org": "Petro"},
    18: {"type": "regular",       "org": "Renzo"},
    19: {"type": "regular",       "org": "Cala"},
    20: {"type": "regular",       "org": None},      # Misa: evento público, sin organizer
    21: {"type": "monthly_event", "org": "Winnie", "vote_avg": 8.0},
    22: {"type": "regular",       "org": "Lucky"},
    23: {"type": "extended",      "org": None},      # Extra: Chinito
    25: {"skip": True},                              # Evento Nacho: ???? (placeholder)
}


# ── Vote helpers ───────────────────────────────────────────────────────────────

def _build_voter_hash(voter_id: str, event_id: str) -> str:
    return hashlib.sha256(f"{voter_id}:{event_id}".encode()).hexdigest()


def _encrypt_payload(fernet: Fernet, fun: int, cost: int, originality: int) -> str:
    raw = json.dumps({"fun": fun, "cost": cost, "originality": originality}).encode()
    return fernet.encrypt(raw).decode()


def build_votes_for_event(event_id: str, attendee_ids: list[str], vote_avg: float,
                           fernet: Fernet) -> list[dict]:
    """
    Generate one synthetic vote per attendee so that:
        ceil(mean(scores) * 10) == organizer_bonus_from_excel

    vote_avg 7.0 → fun=cost=originality=7  → score=7.0 → ceil(7.0*10)=70
    vote_avg 8.0 → fun=cost=originality=8  → score=8.0 → ceil(8.0*10)=80
    """
    v = round(vote_avg)          # 7 or 8 — clean integer
    fun = cost = originality = v
    records = []
    for voter_id in attendee_ids:
        records.append({
            "id": str(uuid.uuid4()),
            "eventId": event_id,
            "voterIdHash": _build_voter_hash(voter_id, event_id),
            "encryptedPayload": _encrypt_payload(fernet, fun, cost, originality),
        })
    return records


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel not found at {EXCEL_PATH}")
        raise SystemExit(1)

    print("\n=== Cebollitas data loader ===\n")

    # 1. Env keys
    print("[1/5] Ensuring .env keys...")
    encryption_key = ensure_env_keys()
    fernet = Fernet(encryption_key.encode())

    # 2. Users
    print("\n[2/5] Building users...")
    allowed_users = [
        {"id": info["id"], "name": info["name"], "email": info["email"]}
        for info in USERS_BY_NICK.values()
    ]
    for u in allowed_users:
        print(f"  {u['name']:<30} {u['id']}")

    # 3. Read Excel
    print("\n[3/5] Reading Excel...")
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        raise SystemExit(1)

    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["aca se ve el descenso (2026)"]

    now_iso = datetime.now(timezone.utc).isoformat()
    events:  list[dict] = []
    votes:   list[dict] = []

    for row_num, meta in sorted(EVENT_META.items()):
        if meta.get("skip"):
            continue

        fecha   = ws.cell(row_num, 3).value
        title   = ws.cell(row_num, 4).value or ""
        raw_obs = ws.cell(row_num, 17).value or ""

        # date: must be a datetime (all included rows have proper dates)
        if not isinstance(fecha, datetime):
            print(f"  SKIP row {row_num}: unexpected date format ({fecha!r})")
            continue

        date_str    = fecha.date().isoformat()
        description = str(raw_obs).strip()[:1000] or title  # fallback to title if empty

        event_type   = meta["type"]
        org_nick     = meta.get("org")

        # Attendees: columns E..P with numeric value > 0
        attendee_nicks = [
            nick for i, nick in enumerate(NICK_ORDER)
            if isinstance(ws.cell(row_num, 5 + i).value, (int, float))
            and ws.cell(row_num, 5 + i).value > 0
        ]
        attendee_ids = [USERS_BY_NICK[n]["id"] for n in attendee_nicks]

        organizer_id = USERS_BY_NICK[org_nick]["id"] if org_nick else None
        creator_id   = organizer_id or LUCAS_EXISTING_ID  # loader user as fallback

        event_id = str(uuid.uuid4())
        event = {
            "id":          event_id,
            "title":       title,
            "description": description,
            "date":        date_str,
            "eventType":   event_type,
            "attendeeIds": attendee_ids,
            "organizerId": organizer_id,
            "creatorId":   creator_id,
            "createdAt":   now_iso,
            "updatedAt":   now_iso,
        }
        events.append(event)
        print(f"  Row {row_num:>2} | {date_str} | {event_type:<14} | org={org_nick or '—':<7} "
              f"| n={len(attendee_ids)} | {title[:45]}")

        # Synthetic votes for monthly_event
        if event_type == "monthly_event":
            vote_avg = meta["vote_avg"]
            event_votes = build_votes_for_event(event_id, attendee_ids, vote_avg, fernet)
            votes.extend(event_votes)
            score = round(vote_avg)
            print(f"         → {len(event_votes)} votes (fun={score} cost={score} orig={score})"
                  f" → organizer bonus = {math.ceil(vote_avg * 10)} pts")

    # 4. Update monthly_assignments: fix Mateo's name
    print("\n[4/5] Updating monthly_assignments.json...")
    monthly_path = DATA_DIR / "monthly_assignments.json"
    assignments  = json.loads(monthly_path.read_text(encoding="utf-8"))
    for month_key, name in assignments.items():
        if name == "Mateo Winnie Pooh":
            assignments[month_key] = "Mateo Lozano"
            print(f"  Month {month_key}: 'Mateo Winnie Pooh' → 'Mateo Lozano'")
    monthly_path.write_text(json.dumps(assignments, ensure_ascii=False, indent=2), encoding="utf-8")

    # 5. Write JSON files
    print("\n[5/5] Writing JSON files...")

    (DATA_DIR / "allowed_users.json").write_text(
        json.dumps(allowed_users, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  allowed_users.json  → {len(allowed_users)} users")

    (DATA_DIR / "events.json").write_text(
        json.dumps(events, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  events.json         → {len(events)} events")

    (DATA_DIR / "votes.json").write_text(
        json.dumps(votes, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  votes.json          → {len(votes)} vote records")

    print("\nDone.\n")


if __name__ == "__main__":
    main()
