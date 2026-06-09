"""load_cebollitas_excel.py — One-time loader: ESTADISTICAS sheet → JSON.

Parses the "PARTIDOS DE CEBOLLITAS" table (team matches) from the source Excel
and rewrites api/data/cebollitas_matches.json, seeded into the
`cebollitas_matches` collection on boot.

Each match is a 5-a-side game: two teams of 5, a winner, a figura (free text)
and an organizer (mapped to an allowed_user). Teams/figura are free text; only
the organizer links to the DB.

Run from the api/ directory:
    venv/Scripts/python.exe scripts/load_cebollitas_excel.py
"""

import json
import uuid
from datetime import datetime
from pathlib import Path

API_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = API_DIR / "data"
EXCEL_PATH = Path(r"D:\Descargas\Partiditos de la gente.xlsx")
SHEET = "ESTADISTICAS"

DEFAULT_CANCHA = 5
DEFAULT_FORMATION = "2-2"

WINNER_BY_LABEL = {"equipo 1": "team1", "equipo 2": "team2", "empate": "draw"}

# Organizer nickname → canonical full name (resolved to userId via allowed_users).
ORGANIZER_BY_NICK = {
    "nacho": "Ignacio Calvo",
    "petro": "Martin Petrone",
    "cala": "Santiago Cala",
}


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _team_players(cells: list) -> list[str]:
    """Collect player names from a column block; the first row may cram them all
    into one cell separated by newlines."""
    names: list[str] = []
    for value in cells:
        text = _clean(value)
        if not text:
            continue
        names.extend(part.strip() for part in text.split("\n") if part.strip())
    return names


def _organizer_id(raw: str | None, id_by_name: dict[str, str]) -> str | None:
    if not raw:
        return None
    nick = raw.strip().rstrip("?").strip().lower()
    name = ORGANIZER_BY_NICK.get(nick)
    if name is None:
        print(f"  WARN: unmapped organizer {raw!r}")
        return None
    return id_by_name.get(name)


def main() -> None:
    if not EXCEL_PATH.exists():
        raise SystemExit(f"ERROR: Excel not found at {EXCEL_PATH}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    users = json.loads((DATA_DIR / "allowed_users.json").read_text(encoding="utf-8"))
    id_by_name = {user["name"]: user["id"] for user in users}

    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[SHEET]

    # Match-start rows = rows (>=17) where column A holds a date.
    starts = [r for r in range(17, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, datetime)]

    now_iso = datetime.now(datetime.now().astimezone().tzinfo).isoformat()
    matches: list[dict] = []
    for i, start in enumerate(starts):
        end = starts[i + 1] - 1 if i + 1 < len(starts) else ws.max_row
        block = range(start, end + 1)

        date_str = ws.cell(start, 1).value.date().isoformat()
        team1 = _team_players([ws.cell(r, 2).value for r in block])
        team2 = _team_players([ws.cell(r, 3).value for r in block])
        winner = WINNER_BY_LABEL.get((_clean(ws.cell(start, 4).value) or "").lower())
        figura = _clean(ws.cell(start, 5).value)
        organizer_id = _organizer_id(_clean(ws.cell(start, 6).value), id_by_name)

        # Pad/trim to the cancha size so the seed stays consistent.
        for label, team in (("Equipo 1", team1), ("Equipo 2", team2)):
            if len(team) != DEFAULT_CANCHA:
                print(f"  WARN row {start} {label}: {len(team)} players (expected {DEFAULT_CANCHA})")
        team1 = (team1 + ["?"] * DEFAULT_CANCHA)[:DEFAULT_CANCHA]
        team2 = (team2 + ["?"] * DEFAULT_CANCHA)[:DEFAULT_CANCHA]

        if winner is None or organizer_id is None:
            print(f"  SKIP row {start}: winner={winner!r} organizer={organizer_id!r}")
            continue

        matches.append(
            {
                "id": str(uuid.uuid5(uuid.NAMESPACE_DNS, f"cebollitas:teammatch:{start}:{date_str}")),
                "date": date_str,
                "cancha": DEFAULT_CANCHA,
                "team1": {"formation": DEFAULT_FORMATION, "players": team1},
                "team2": {"formation": DEFAULT_FORMATION, "players": team2},
                "winner": winner,
                "figura": figura,
                "organizerId": organizer_id,
                "creatorId": id_by_name.get("Lucas Mateo Sacks", ""),
                "createdAt": now_iso,
                "updatedAt": now_iso,
            }
        )

    out_path = DATA_DIR / "cebollitas_matches.json"
    out_path.write_text(json.dumps(matches, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Done. {len(matches)} cebollitas matches written to {out_path}.")


if __name__ == "__main__":
    main()
